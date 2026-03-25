from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def build_excel_report(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    export_df = df.copy()

    ordered_columns = [
        "proforma",
        "data",
        "cliente",
        "cnpj",
        "contrato_po",
        "po_rm_pedido",
        "ordem_venda",
        "nf",
        "data_nf",
        "mes_contabil",
        "observacoes",
        "bu",
        "details",
        "well_project",
        "valor_bruto_brl",
        "valor_bruto_usd",
        "valor_faturado_brl",
        "valor_liquido_brl",
        "impostos",
        "percentual_impostos",
        "status",
        "comentarios_variacao",
        "origem_aba",
        "origem_linha",
        "tem_dado_incompleto",
        "valor_total_considerado",
    ]
    export_df = export_df[[column for column in ordered_columns if column in export_df.columns]]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Relatorio", index=False)

        resumo_linhas = [
            {"Indicador": "Total de registros", "Valor": len(export_df)},
            {
                "Indicador": "Total bruto BRL",
                "Valor": export_df["valor_bruto_brl"].sum(min_count=1) if "valor_bruto_brl" in export_df else 0,
            },
            {
                "Indicador": "Total líquido BRL",
                "Valor": export_df["valor_liquido_brl"].sum(min_count=1) if "valor_liquido_brl" in export_df else 0,
            },
            {
                "Indicador": "Total impostos",
                "Valor": export_df["impostos"].sum(min_count=1) if "impostos" in export_df else 0,
            },
            {
                "Indicador": "Pendências",
                "Valor": int(export_df["status"].fillna("").str.contains("PENDENTE|UNBILLED", case=False).sum()) if "status" in export_df else 0,
            },
        ]
        pd.DataFrame(resumo_linhas).to_excel(writer, sheet_name="Resumo", index=False)

        if {"bu", "valor_total_considerado"}.issubset(export_df.columns):
            resumo_bu = (
                export_df.groupby("bu", dropna=False)["valor_total_considerado"]
                .sum(min_count=1)
                .fillna(0)
                .reset_index()
                .sort_values("valor_total_considerado", ascending=False)
            )
            resumo_bu.to_excel(writer, sheet_name="Resumo_BU", index=False)

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font

            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    if len(value) > max_length:
                        max_length = len(value)
                worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)

            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    header_value = worksheet.cell(1, cell.column).value
                    if header_value in {
                        "valor_bruto_brl",
                        "valor_bruto_usd",
                        "valor_faturado_brl",
                        "valor_liquido_brl",
                        "impostos",
                        "valor_total_considerado",
                    }:
                        cell.number_format = 'R$ #,##0.00'
                    elif header_value == "percentual_impostos":
                        cell.number_format = "0.00%"
                    elif header_value in {"data", "data_nf"}:
                        cell.number_format = "dd/mm/yyyy"

    return output.getvalue()
