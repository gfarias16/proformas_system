from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

SOURCE_FILE = Path("PROFORMAS 2026 (1).xlsx")
OUTPUT_CSV = Path("base_normalizada.csv")
OUTPUT_XLSX = Path("base_normalizada.xlsx")
SOURCE_SHEETS = [
    "RCAL",
    "SCAL",
    "GEOCHEMISTRY",
    "GEOLOGY",
    "STORAGE-LOGISTC",
    "PRODUCT SALES",
]

HEADER_ALIASES = {
    "PROFORMA": "proforma",
    "DATA": "data",
    "CLIENTE": "cliente",
    "CNPJ": "cnpj",
    "CONTRATO/ PO": "contrato_po",
    "PO/RM/PEDIDO": "po_rm_pedido",
    "ORDEM DE VENDA": "ordem_venda",
    "N.F.": "nf",
    "N.F. ": "nf",
    "DATA N.F.": "data_nf",
    "MÊS CONTABIL": "mes_contabil",
    "OBSERVAÇÕES": "observacoes",
    "BU": "bu",
    "DETAILS": "details",
    "WELL/PROJECT": "well_project",
    "VALOR BRUTO BRL - PF": "valor_bruto_brl",
    "VALOR BRUTO BRL  - PF": "valor_bruto_brl",
    "VALOR BRUTO USD": "valor_bruto_usd",
    "VALOR FATURADO BRL": "valor_faturado_brl",
    "VALOR LÍQUIDO BRL": "valor_liquido_brl",
    "IMPOSTOS": "impostos",
    "%": "percentual_impostos",
    "STATUS": "status",
    "VARIAÇÃO NO FATURAMENTO": "variacao_faturamento",
    "COMENTÁRIOS SOBRE A VARIAÇÃO:": "comentarios_variacao",
    "CL": "cl",
    "ENVIADO AO CLIENTE EM:": "enviado_cliente_em",
}

NUMERIC_COLUMNS = [
    "valor_bruto_brl",
    "valor_bruto_usd",
    "valor_faturado_brl",
    "valor_liquido_brl",
    "impostos",
    "percentual_impostos",
]
DATE_COLUMNS = ["data", "data_nf", "enviado_cliente_em"]


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = re.sub(r"\s+", " ", text)
    text = text.replace(" .", ".")
    return HEADER_ALIASES.get(text, text.lower())


def is_valid_proforma(value: object) -> bool:
    return bool(re.match(r"^\d{4}\.\d+", str(value).strip())) if value else False


def sanitize_value(value: object) -> object:
    if isinstance(value, str):
        value = value.strip()
        if not value or value in {"#DIV/0!", "#REF!", "#NAME?", "#N/A"}:
            return None
    return value


def parse_mixed_date(value: object) -> pd.Timestamp:
    if value in (None, "", 0):
        return pd.NaT
    if isinstance(value, pd.Timestamp):
        return value
    try:
        return pd.to_datetime(value, errors="coerce")
    except (TypeError, ValueError):
        return pd.NaT


def find_header_row(worksheet) -> int:
    for row_index in range(1, 120):
        if str(worksheet.cell(row_index, 1).value).strip().upper() == "PROFORMA":
            return row_index
    raise ValueError(f"Cabeçalho não encontrado na aba {worksheet.title}.")


def extract_sheet_rows(worksheet) -> list[dict]:
    header_row = find_header_row(worksheet)
    raw_headers = [worksheet.cell(header_row, col).value for col in range(1, 25)]
    headers = [normalize_header(header) for header in raw_headers]
    records: list[dict] = []
    blank_run = 0

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        values = [sanitize_value(worksheet.cell(row_index, col).value) for col in range(1, 25)]
        proforma = values[0]

        if is_valid_proforma(proforma):
            blank_run = 0
            record = {
                "origem_aba": worksheet.title,
                "origem_linha": row_index,
            }
            for header, value in zip(headers, values):
                if not header:
                    continue
                record[header] = value
            records.append(record)
            continue

        if all(value is None for value in values[:16]):
            blank_run += 1
            if blank_run >= 200:
                break
        else:
            blank_run = 0

    return records


def build_dataframe() -> pd.DataFrame:
    workbook = load_workbook(SOURCE_FILE, data_only=True, read_only=False)
    all_records: list[dict] = []

    for sheet_name in SOURCE_SHEETS:
        worksheet = workbook[sheet_name]
        all_records.extend(extract_sheet_rows(worksheet))

    df = pd.DataFrame(all_records)
    if df.empty:
        raise ValueError("Nenhum registro válido foi encontrado nas abas operacionais.")

    for column in NUMERIC_COLUMNS:
        if column in df.columns:
            df[column] = pd.to_numeric(df[column], errors="coerce")

    for column in DATE_COLUMNS:
        if column in df.columns:
            df[column] = df[column].apply(parse_mixed_date)

    if "mes_contabil" in df.columns:
        df["mes_contabil"] = df["mes_contabil"].astype("string").str.strip()

    df["status"] = df.get("status", pd.Series(dtype="object")).fillna("PENDENTE")
    df["registro_id"] = (
        df["origem_aba"].astype(str)
        + "::"
        + df["origem_linha"].astype(str)
        + "::"
        + df["proforma"].astype(str)
    )
    df["tem_dado_incompleto"] = df[["cliente", "mes_contabil", "bu"]].isna().any(axis=1)
    valor_liquido = df["valor_liquido_brl"] if "valor_liquido_brl" in df.columns else pd.Series(index=df.index, dtype="float64")
    valor_faturado = df["valor_faturado_brl"] if "valor_faturado_brl" in df.columns else pd.Series(index=df.index, dtype="float64")
    valor_bruto = df["valor_bruto_brl"] if "valor_bruto_brl" in df.columns else pd.Series(index=df.index, dtype="float64")
    df["valor_total_considerado"] = valor_liquido.fillna(valor_faturado).fillna(valor_bruto)

    preferred_order = [
        "registro_id",
        "origem_aba",
        "origem_linha",
        "proforma",
        "data",
        "cliente",
        "cnpj",
        "contrato_po",
        "po_rm_pedido",
        "ordem_venda",
        "nf",
        "data_nf",
        "mes_contabil",
        "observacoes",
        "bu",
        "details",
        "well_project",
        "valor_bruto_brl",
        "valor_bruto_usd",
        "valor_faturado_brl",
        "valor_liquido_brl",
        "impostos",
        "percentual_impostos",
        "status",
        "comentarios_variacao",
        "cl",
        "enviado_cliente_em",
        "tem_dado_incompleto",
        "valor_total_considerado",
    ]
    ordered_columns = [column for column in preferred_order if column in df.columns]
    remaining_columns = [column for column in df.columns if column not in ordered_columns]
    return df[ordered_columns + remaining_columns].sort_values(
        by=["mes_contabil", "bu", "cliente", "proforma"],
        na_position="last",
    )


def export_files(df: pd.DataFrame) -> None:
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    df.to_excel(OUTPUT_XLSX, index=False)


def print_summary(df: pd.DataFrame) -> None:
    resumo_bu = (
        df.groupby("bu", dropna=False)["valor_total_considerado"]
        .sum(min_count=1)
        .sort_values(ascending=False)
        .fillna(0)
    )
    print(f"Base normalizada gerada com {len(df)} registros.")
    print(f"Arquivos: {OUTPUT_CSV} e {OUTPUT_XLSX}")
    print("Resumo por BU:")
    print(resumo_bu.to_string())


def main() -> None:
    df = build_dataframe()
    export_files(df)
    print_summary(df)


if __name__ == "__main__":
    main()
